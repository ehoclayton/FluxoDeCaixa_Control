import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QLabel, QLineEdit, QPushButton, QComboBox, QMessageBox, QGridLayout
from PyQt5.QtGui import QFont
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

class FluxoDeCaixaApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Fluxo de Caixa")
        self.setGeometry(100, 100, 500, 50)
        self.setStyleSheet(
            """
            QMainWindow {
                background: #6fc1ff; /* Cor de fundo azul céu */
            }
            QWidget {
                color: black;
            }
            """
        )
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.central_widget.setStyleSheet("background: transparent; color: black;")

        # Crie um layout de grade para organizar os widgets
        grid = QGridLayout(self.central_widget)

        self.initUI(grid)

        # Carregue as categorias no início
        self.categorias_existentes = []  # Inicialize a lista de categorias
        self.carregar_categorias()

    def initUI(self, grid):
        font = QFont("Arial", 12)

        # Rótulos
        self.tipo_label = QLabel("Escolha uma transição:")
        self.tipo_label.setFont(font)

        self.descricao_label = QLabel("Descrição:")
        self.descricao_label.setFont(font)

        self.valor_label = QLabel("Valor:")
        self.valor_label.setFont(font)

        self.rs_label = QLabel("R$:")
        self.rs_label.setFont(font)

        self.categoria_label = QLabel("Categoria:")
        self.categoria_label.setFont(font)

        self.add_del_categoria_label = QLabel("Atualizar Categorias:")
        self.add_del_categoria_label.setFont(font)

        # Entradas de texto
        self.descricao_entry = QLineEdit(self.central_widget)
        self.descricao_entry.setFont(font)
        self.descricao_entry.setStyleSheet(
            """
            background-color: white; /* Cor de fundo branca */
            color: black; /* Cor do texto preto */
            border-radius: 5px; /* Borda arredondada */
            """
        )

        self.valor_entry = QLineEdit(self.central_widget)
        self.valor_entry.setFont(font)
        self.valor_entry.setStyleSheet(
            """
            background-color: white; /* Cor de fundo branca */
            color: black; /* Cor do texto preto */
            border-radius: 5px; /* Borda arredondada */
            """
        )

        # Dropdown para Tipo de Transação
        self.tipo_dropdown = QComboBox(self.central_widget)
        self.tipo_dropdown.addItems(["Entrada", "Saída"])
        self.tipo_dropdown.setFont(font)
        self.tipo_dropdown.setStyleSheet(
            """
            QComboBox {
                background-color: #F5F5F5; /* Cor de fundo cinza claro */
                border: 1px solid #CCCCCC; /* Borda cinza claro */
                border-radius: 5px; /* Borda arredondada */
                color: black; /* Cor do texto no botão */
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 20px; /* Largura do botão de drop-down */
                border-left: 1px solid #CCCCCC; /* Borda esquerda cinza claro */
                border-top-right-radius: 5px; /* Borda arredondada superior direita */
                border-bottom-right-radius: 5px; /* Borda arredondada inferior direita */
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #F5F5F5, stop:1 #E0E0E0); /* Gradiente para o botão de drop-down */
            }
            QComboBox::down-arrow {
                image: url(down_arrow.png); /* Ícone do botão de drop-down (substitua 'down_arrow.png' pelo seu ícone) */
            }
            QComboBox QAbstractItemView {
                background-color: #F5F5F5; /* Cor de fundo do menu drop-down */
                color: black; /* Cor do texto das opções dentro do menu drop-down */
            }
            """
        )

        # Dropdown para Categoria
        self.categoria_dropdown = QComboBox(self.central_widget)
        self.categoria_dropdown.setFont(font)
        self.categoria_dropdown.setStyleSheet(
            """
            QComboBox {
                background-color: white; /* Cor de fundo branca */
                color: black; /* Cor do texto preto */
                border-radius: 5px; /* Borda arredondada */
            }
            QComboBox::down-arrow {
                image: url(down_arrow.png); /* Ícone da seta (substitua 'down_arrow.png' pelo seu ícone) */
                width: 12px; /* Largura da seta */
                height: 12px; /* Altura da seta */
            }
            QComboBox QAbstractItemView {
                background-color: white; /* Cor de fundo branca */
                color: black; /* Cor do texto preto */
                border: 1px solid #CCCCCC; /* Borda cinza claro */
            }
            """
        )

        # Entrada de texto para nova categoria
        self.nova_categoria_entry = QLineEdit(self.central_widget)
        self.nova_categoria_entry.setFont(font)
        self.nova_categoria_entry.setStyleSheet(
            """
            background-color: white; /* Cor de fundo branca */
            color: black; /* Cor do texto preto */
            border-radius: 5px; /* Borda arredondada */
            """
        )

        # Botão para adicionar nova categoria
        self.adicionar_categoria_button = QPushButton("+", self.central_widget)
        self.adicionar_categoria_button.setFont(font)
        self.adicionar_categoria_button.setStyleSheet(
            """
            background-color: #4CAF50; /* Cor de fundo verde */
            border: none;
            color: white; /* Cor do texto branco */
            padding: 5px 10px; /* Reduzir o espaço interno */
            border-radius: 5px; /* Borda arredondada */
            """
        )
        self.adicionar_categoria_button.clicked.connect(self.adicionar_categoria)

        # Botão para remover categoria
        self.remover_categoria_button = QPushButton("-", self.central_widget)
        self.remover_categoria_button.setFont(font)
        self.remover_categoria_button.setStyleSheet(
            """
            background-color: #FF3333; /* Cor de fundo vermelha */
            border: none;
            color: white; /* Cor do texto branco */
            padding: 5px 10px; /* Reduzir o espaço interno */
            border-radius: 5px; /* Borda arredondada */
            """
        )
        self.remover_categoria_button.clicked.connect(self.remover_categoria)

        # Rótulo de status
        self.status_label = QLabel("", self.central_widget)
        self.status_label.setFont(font)
        self.status_label.setStyleSheet("color: green;")

        # Botões
        self.registrar_button = QPushButton("Registrar", self.central_widget)
        self.registrar_button.setFont(font)
        self.registrar_button.setStyleSheet(
            """
            background-color: #2196F3; /* Cor de fundo azul */
            border: none;
            color: white; /* Cor do texto branco */
            padding: 5px 10px; /* Reduzir o espaço interno */
            border-radius: 5px; /* Borda arredondada */
            """
        )
        self.registrar_button.clicked.connect(self.registrar_transacao)

        self.fechar_button = QPushButton("Fechar", self.central_widget)
        self.fechar_button.setFont(font)
        self.fechar_button.setStyleSheet(
            """
            background-color: #666; /* Cor de fundo cinza */
            border: none;
            color: white; /* Cor do texto branco */
            padding: 5px 10px; /* Reduzir o espaço interno */
            border-radius: 5px; /* Borda arredondada */
            """
        )
        self.fechar_button.clicked.connect(self.close)

        # Posicionamento dos widgets na grade
        grid.addWidget(self.tipo_label, 0, 0)
        grid.addWidget(self.tipo_dropdown, 0, 1, 1, 3)

        grid.addWidget(self.descricao_label, 1, 0)
        grid.addWidget(self.descricao_entry, 1, 1, 1, 3)

        grid.addWidget(self.valor_label, 2, 0)
        grid.addWidget(self.rs_label, 2, 1)
        grid.addWidget(self.valor_entry, 2, 2, 1, 2)

        grid.addWidget(self.categoria_label, 3, 0)
        grid.addWidget(self.categoria_dropdown, 3, 1, 1, 2)

        grid.addWidget(self.add_del_categoria_label, 4, 0)
        grid.addWidget(self.nova_categoria_entry, 4, 1, 1, 2)
        grid.addWidget(self.adicionar_categoria_button, 4, 3, 1, 1)
        grid.addWidget(self.remover_categoria_button, 3, 3, 1, 1)

        grid.addWidget(self.registrar_button, 5, 0, 1, 4)
        grid.addWidget(self.fechar_button, 6, 0, 1, 4)
        grid.addWidget(self.status_label, 7, 1)

    def carregar_categorias(self):
        try:
            with open('categorias.txt', 'r') as file:
                categorias = file.read().splitlines()
                if categorias:
                    self.categoria_dropdown.addItems(categorias)
                    self.categorias_existentes = categorias
        except FileNotFoundError:
            self.categorias_existentes = []

    def salvar_categorias(self):
        # Atualize a variável de instância com as categorias existentes
        self.categorias_existentes = [self.categoria_dropdown.itemText(i) for i in range(self.categoria_dropdown.count())]
        with open('categorias.txt', 'w') as file:
            for categoria in self.categorias_existentes:
                file.write(categoria + '\n')

    def atualizar_categorias(self):
        # Limpar o dropdown
        self.categoria_dropdown.clear()
        # Adicionar as categorias existentes de volta ao dropdown
        self.categoria_dropdown.addItems(self.categorias_existentes)

    def adicionar_categoria(self):
        print("Botão Adicionar Categoria clicado")
        nova_categoria = self.nova_categoria_entry.text()
        if nova_categoria and nova_categoria not in [self.categoria_dropdown.itemText(i) for i in range(self.categoria_dropdown.count())]:
            self.categoria_dropdown.addItem(nova_categoria)
            self.salvar_categorias()
            self.nova_categoria_entry.clear()
            QMessageBox.information(self, "Sucesso", "Categoria adicionada com sucesso!")

    def remover_categoria(self):
        categoria_selecionada = self.categoria_dropdown.currentText()
        if categoria_selecionada in self.categorias_existentes:
            self.categorias_existentes.remove(categoria_selecionada)  # Remova a categoria da lista
            self.salvar_categorias()  # Salve as categorias atualizadas no arquivo de texto
            index = self.categoria_dropdown.findText(categoria_selecionada)
            if index >= 0:
                self.categoria_dropdown.removeItem(index)
                self.salvar_categorias()
            print("Categoria removida:", categoria_selecionada)  # Adicione esta linha para depuração
            QMessageBox.information(self, "Sucesso", "Categoria removida com sucesso!")
        else:
            QMessageBox.warning(self, "Aviso", "Categoria não encontrada.")




    def registrar_transacao(self):
        tipo_transacao = self.tipo_dropdown.currentText()
        descricao = self.descricao_entry.text()
        valor_str = self.valor_entry.text().replace(",", ".")
        categoria = self.categoria_dropdown.currentText()

        valor = 0.0

        if not tipo_transacao or not descricao or not valor_str or not categoria:
            QMessageBox.critical(self, "Erro", "Preencha todos os campos antes de registrar!")
            return

        try:
            valor = float(valor_str)
        except ValueError:
            QMessageBox.critical(self, "Erro", "Valor inválido! Certifique-se de usar um número válido.")
            return

        if tipo_transacao == "Saída":
            valor = -valor

        try:
            workbook = openpyxl.load_workbook('fluxo_de_caixa.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Data", "Tipo", "Descrição", "Valor", "Categoria"])

        next_row = sheet.max_row + 1

        data_hora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=next_row, column=1, value=data_hora)
        sheet.cell(row=next_row, column=2, value=tipo_transacao)
        sheet.cell(row=next_row, column=3, value=descricao)
        sheet.cell(row=next_row, column=4, value=valor)
        sheet.cell(row=next_row, column=5, value=categoria)

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

        workbook.save('fluxo_de_caixa.xlsx')
        QMessageBox.information(self, "Sucesso", "Transação registrada com sucesso!")

        self.descricao_entry.clear()
        self.valor_entry.clear()
        self.categoria_dropdown.setCurrentIndex(-1)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FluxoDeCaixaApp()
    window.show()
    sys.exit(app.exec_())